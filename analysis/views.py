from django.shortcuts import render
from students.wordpress import WordPress
from .models import Member, Cursus
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import user_passes_test
from .forms import XlsxUpload, Confirm, ChooseCursus
from io import BytesIO
from openpyxl import load_workbook
from django.db.models import Count, Avg
import yaml


@user_passes_test(lambda u: u.is_superuser)
def load_members(request):
    if request.method == "POST":
        form = Confirm(request.POST)

        if form.is_valid():
            Member.objects.all().delete()

            props, data = WordPress.get_students_data(board=True)

            for mdata in data:
                m = Member()
                m.load_from_csv(props, mdata)
                m.save()

            return render(request, 'base.html', {
                'message' : 'All members imported to analysis database!'
            })
    else:
        form = Confirm()

    return render(request, "confirm.html", {
        "subject" : "importing members, delete existing",
        "form" : form
    })

@user_passes_test(lambda u: u.is_superuser)
def upload_subscriptions(request):
    if request.method == 'POST':
        form = XlsxUpload(request.POST, request.FILES)
        if form.is_valid():
            file_in_memory = request.FILES['file'].read()
            wb = load_workbook(filename=BytesIO(file_in_memory), read_only=True)
            Cursus.objects.all().delete()
            for sheet in wb.worksheets:
                c = Cursus(name=sheet.title)
                c.save()
                iterrows = sheet.iter_rows()
                next(iterrows) #skip first row
                for row in iterrows:
                    data = list(row)
                    if data[0].value is None:
                        continue
                    try:
                        m = Member.objects.get(first_name=data[0].value.lower().strip(), last_name=data[1].value.lower().strip())
                    except Member.DoesNotExist:
                        print("Could not find {} {} for {}".format(data[0].value.lower(), data[1].value.lower(), c))
                        continue
                    m.subscriptions.add(c)
                    m.save()
            return render(request, 'base.html', {
                'message' : 'subscriptions imported'
            })
    else:
        form = XlsxUpload()
    return render(request, 'upload.html', {
        'form' : form
    })

@staff_member_required
def stats(request):
    #TODO: allow multiple courses to be selected
    if request.method == "POST":
        form = ChooseCursus(request.POST)
        if form.is_valid():
            base_set = Member.objects.annotate(cursus_count=Count('subscriptions')).filter(subscriptions=form.cleaned_data['cursus'])
            name = str(form.cleaned_data['cursus'])
        else:
            base_set = Member.objects.annotate(cursus_count=Count('subscriptions')).filter(cursus_count__gt=0)
            name = 'all'
    else:
        form = ChooseCursus()
        base_set = Member.objects.annotate(cursus_count=Count('subscriptions')).filter(cursus_count__gt=0)
        name = 'all'

    total_ammount = base_set.count()
    total_ammount_student = base_set.filter(student=True).count()
    total_ammount_tue = base_set.filter(institute=0).count()
    total_ammount_fon = base_set.filter(institute=1).count()
    total_ammount_student_eindhoven = base_set.filter(institute__lt=4).count()

    percentage_student = round(total_ammount_student / total_ammount * 100, 2)
    percentage_tue = round(total_ammount_tue / total_ammount_student * 100, 2)
    percentage_fon = round(total_ammount_fon / total_ammount_student * 100, 2)
    percentage_student_eindhoven_of_total = round(total_ammount_student_eindhoven / total_ammount * 100, 2)
    percentage_student_eindhoven_of_students = round(total_ammount_student_eindhoven / total_ammount_student * 100, 2)

    men_ammount = base_set.filter(gender='M').count()
    woman_ammount = base_set.filter(gender='F').count()

    percentage_men = round(men_ammount / total_ammount * 100, 2)
    percentage_woman = round(woman_ammount / total_ammount * 100, 2)

    avg_cursus_per_member = round(base_set.aggregate(Avg('cursus_count'))['cursus_count__avg'], 2)


    data = {
        'total ammount' : total_ammount,
        'total ammount student' : total_ammount_student,
        'total ammount tue of student' : total_ammount_tue,
        'total ammount fontys of student' : total_ammount_fon,
        'percentage student' : percentage_student,
        'percentage tue' : percentage_tue,
        'percentage fontys' : percentage_fon,
        'men ammount' : men_ammount,
        'woman ammount' : woman_ammount,
        'percentage men' : percentage_men,
        'percentage woman' : percentage_woman,
        'percentage student eindhoven of total' : percentage_student_eindhoven_of_total,
        'percentage student eindhoven of students' : percentage_student_eindhoven_of_students,
        'average cursus per member' : avg_cursus_per_member,
    }

    with open('stats_{}.yaml'.format(name), 'w') as stream:
        yaml.dump(data, stream, default_flow_style=False)

    return render(request, 'stats.html', {
        'data' : data,
        'subject' : name,
        'form' : form
    })

@staff_member_required
def list_all(request):
    return render(request, 'analysis_list_all.html', {
        'cursussen' : Cursus.objects.all()
    })